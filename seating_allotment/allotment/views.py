from django.db.models import *
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import *
from .utils import generate_auto_mixed_pg_seating
import openpyxl
from django.db import transaction
from django.template.loader import render_to_string
from weasyprint import HTML
from django.http import HttpResponse
from django.utils.timezone import now
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
import openpyxl
from openpyxl import Workbook



@never_cache
def staff_logout(request):
    logout(request)
    request.session.flush()   # completely destroy session
    response = redirect("login")
    response.delete_cookie("sessionid")
    return response

@never_cache

def staff_login(request):
    if request.method == "POST":
        staff_id = request.POST.get("staff_id")
        password = request.POST.get("password")

        user = authenticate(request, username=staff_id, password=password)

        if user is not None:
            login(request, user)
            return redirect("dashboard")  # change if needed
        else:
            messages.error(request, "Invalid Staff ID or Password")

    return render(request, "allotment/staff_login.html")

@never_cache
@login_required(login_url="login")

def export_master_pdf(request):
    allotments = SeatingAllotment.objects.all().order_by(
        'hall__hall_no',
        'seat_number'
    )

    html_string = render_to_string(
        'allotments/master_list_pdf.html',
        {
            'allotments': allotments,
            'now': now()
        }
    )

    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="Seating_Master_List.pdf"'

    return response



# ==========================================
# DASHBOARD
# ==========================================
@never_cache
@login_required(login_url="login")
def dashboard(request):

    student_count = Student.objects.count()
    subject_count = Subject.objects.count()
    hall_count = ExamHall.objects.count()
    department_count = Department.objects.count()
    exam_count = Exam.objects.count()

    # 🔴 Unallocated Students
    allocated_students = SeatingAllotment.objects.values('student').distinct().count()
    unallocated_students = student_count - allocated_students

    # 🌅 FN / 🌇 AN Exam Summary
    fn_count = Exam.objects.filter(session='FN').count()
    an_count = Exam.objects.filter(session='AN').count()

    allotments = SeatingAllotment.objects.select_related(
        'student',
        'exam',
        'hall',
        'hall__building',
        'exam__subject'
    ).order_by(
        'exam__exam_date',
        'exam__session',
        'hall__hall_no',
        'seat_number'
    )[:200]

    context = {
        'student_count': student_count,
        'subject_count': subject_count,
        'hall_count': hall_count,
        'department_count': department_count,
        'exam_count': exam_count,
        'unallocated_students': unallocated_students,
        'fn_count': fn_count,
        'an_count': an_count,
        'allotments': allotments,
    }

    return render(request, 'allotment/dashboard.html', context)




# ==========================================
# GENERATE SEATING
# ==========================================
@never_cache
@login_required(login_url="login")
def generate_pg_mixed_allotments(request):
    SeatingAllotment.objects.all().delete()

    exams = Exam.objects.order_by('exam_date', 'session')
    halls = list(ExamHall.objects.order_by('hall_no'))

    if not exams.exists():
        messages.error(request, "No exams available to generate seating.")
        return redirect('dashboard')

    exam_groups = {}

    for exam in exams:
        key = (exam.exam_date, exam.session)
        exam_groups.setdefault(key, []).append(exam)

    for (exam_date, session), grouped_exams in exam_groups.items():
        generate_auto_mixed_pg_seating(grouped_exams, halls)

    messages.success(request, "Seating generated successfully.")
    return redirect('dashboard')


# ==========================================
# STUDENT PAGE
# ==========================================
@never_cache
@login_required(login_url="login")
def student_page(request):

    query = request.GET.get("q", "").strip()

    students = Student.objects.select_related("department")

    if query:
        students = students.filter(
            Q(name__icontains=query) |
            Q(reg_no__icontains=query) |
            Q(department__name__icontains=query) |
            Q(semester__icontains=query)
        )

    students = students.order_by("reg_no")

    departments = Department.objects.all()
    edit_student = None

    # ➕ Add Student
    if request.method == "POST" and "add_student" in request.POST:
        Student.objects.create(
            name=request.POST["name"],
            reg_no=request.POST["reg_no"],
            department_id=request.POST["department"],
            semester=request.POST["semester"]
        )
        return redirect("student_page")

    # ✏ Update Student
    if request.method == "POST" and "update_student" in request.POST:
        student = get_object_or_404(Student, id=request.POST["student_id"])
        student.name = request.POST["name"]
        student.reg_no = request.POST["reg_no"]
        student.department_id = request.POST["department"]
        student.semester = request.POST["semester"]
        student.save()
        return redirect("student_page")

    # 📝 Edit Mode
    if request.GET.get("edit"):
        edit_student = get_object_or_404(Student, id=request.GET["edit"])

    # 🗑 Delete
    if request.GET.get("delete"):
        Student.objects.filter(id=request.GET["delete"]).delete()
        return redirect("student_page")

    return render(request, "allotment/student_page.html", {
        "students": students,
        "departments": departments,
        "edit_student": edit_student,
    })


# ==========================================
# HALL PAGE
# ==========================================
@never_cache
@login_required(login_url="login")
def hall_page(request):

    query = request.GET.get("q")

    halls = ExamHall.objects.select_related("building")

    if query:
        halls = halls.filter(
            Q(hall_no__icontains=query) |
            Q(building__name__icontains=query)
        )

    halls = halls.order_by('hall_no')

    buildings = Building.objects.all()
    edit_hall = None

    if request.GET.get('edit'):
        edit_hall = get_object_or_404(ExamHall, id=request.GET.get('edit'))

    if request.GET.get('delete'):
        hall = get_object_or_404(ExamHall, id=request.GET.get('delete'))
        hall.delete()
        return redirect('hall_page')

    if request.method == "POST" and "add_hall" in request.POST:
        ExamHall.objects.create(
            hall_no=request.POST.get("hall_no"),
            total_seats=request.POST.get("capacity"),
            building_id=request.POST.get("building")
        )
        return redirect('hall_page')

    if request.method == "POST" and "update_hall" in request.POST:
        hall = get_object_or_404(ExamHall, id=request.POST.get("hall_id"))
        hall.hall_no = request.POST.get("hall_no")
        hall.total_seats = request.POST.get("capacity")
        hall.building_id = request.POST.get("building")
        hall.save()
        return redirect('hall_page')

    return render(request, "allotment/hall_page.html", {
        "halls": halls,
        "buildings": buildings,
        "edit_hall": edit_hall,
    })

@never_cache
@login_required(login_url="login")
def subjects_page(request):

    # -----------------------------
    # CREATE SUBJECT + EXAM
    # -----------------------------
    if request.method == "POST":

        subject_code = request.POST.get("subject_code").strip()
        subject_name = request.POST.get("subject_name").strip()
        department_id = request.POST.get("department")
        semester = request.POST.get("semester")
        exam_date = request.POST.get("exam_date")
        session = request.POST.get("session")

        department = get_object_or_404(Department, id=department_id)

        # Prevent duplicate crash
        subject, created = Subject.objects.update_or_create(
            subject_code=subject_code,
            defaults={
                "subject_name": subject_name,
                "department": department,
                "semester": semester
            }
        )

        # Create exam only if not already existing
        if exam_date and session:
            Exam.objects.get_or_create(
                subject=subject,
                exam_date=exam_date,
                session=session
            )

        return redirect("subjects")

    # -----------------------------
    # SEARCH FUNCTIONALITY
    # -----------------------------
    query = request.GET.get("search")


    subjects = Subject.objects.select_related("department")

    if query:
        subjects = subjects.filter(
            Q(subject_name__icontains=query) |
            Q(subject_code__icontains=query) |
            Q(department__name__icontains=query) |
            Q(semester__icontains=query)
        )

    subjects = subjects.order_by("subject_code")

    departments = Department.objects.all()

    return render(request, "allotment/Subjects.html", {
    "subjects": subjects,
    "departments": departments,
    "search_query": query
})

@never_cache
@login_required(login_url="login")
def delete_subject(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    subject.delete()
    return redirect("subjects")




# ==========================================
# DEPARTMENT PAGE
# ==========================================
@never_cache
@login_required(login_url="login")
def departments_page(request):

    if request.method == "POST":
        name = request.POST.get("name", "").strip()

        if not name:
            messages.error(request, "Department name is required.")
            return redirect("departments")

        if Department.objects.filter(name__iexact=name).exists():
            messages.error(request, "Department already exists.")
            return redirect("departments")

        Department.objects.create(name=name)
        messages.success(request, "Department added successfully.")
        return redirect("departments")

    search = request.GET.get("search", "").strip()
    departments = Department.objects.all()

    if search:
        departments = departments.filter(
            Q(name__icontains=search)
        )

    return render(request, "allotment/departments.html", {
        "departments": departments,
        "search": search
    })

@never_cache
@login_required(login_url="login")
def delete_department(request, pk):
    department = get_object_or_404(Department, pk=pk)

    if department.students.exists() or department.subjects.exists():
        messages.error(request, "Cannot delete department linked to students or subjects.")
        return redirect("departments")

    department.delete()
    messages.success(request, "Department deleted successfully.")
    return redirect("departments")

@never_cache
@login_required(login_url="login")
def master_upload(request):

    if request.method == "POST" and request.FILES.get("excel_file"):

        file = request.FILES["excel_file"]

        try:
            wb = openpyxl.load_workbook(file)

            with transaction.atomic():

                # =========================
                # Departments
                # =========================
                if "Departments" in wb.sheetnames:
                    sheet = wb["Departments"]
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            Department.objects.get_or_create(
                                name=row[0].strip()
                            )

                # =========================
                # Students
                # =========================
                if "Students" in wb.sheetnames:
                    sheet = wb["Students"]
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        reg_no, name, dept_name, semester = row

                        if not reg_no:
                            continue

                        department = Department.objects.get(
                            name=dept_name.strip()
                        )

                        Student.objects.update_or_create(
                            reg_no=str(reg_no).strip(),
                            defaults={
                                "name": name.strip(),
                                "department": department,
                                "semester": int(semester),
                            }
                        )

                # =========================
                # Subjects
                # =========================
                if "Subjects" in wb.sheetnames:
                    sheet = wb["Subjects"]
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        code, name, dept_name, semester = row

                        department = Department.objects.get(
                            name=dept_name.strip()
                        )

                        Subject.objects.update_or_create(
                            subject_code=code.strip(),
                            defaults={
                                "subject_name": name.strip(),
                                "department": department,
                                "semester": int(semester),
                            }
                        )

                # =========================
                # Exams
                # =========================
                if "Exams" in wb.sheetnames:
                    sheet = wb["Exams"]
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        subject_code, exam_date, session = row

                        subject = Subject.objects.get(
                            subject_code=subject_code.strip()
                        )

                        Exam.objects.get_or_create(
                            subject=subject,
                            exam_date=exam_date,
                            session=session
                        )

                # =========================
                # Buildings
                # =========================
                if "Buildings" in wb.sheetnames:
                    sheet = wb["Buildings"]
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            Building.objects.get_or_create(
                                name=row[0].strip()
                            )

                # =========================
                # Halls
                # =========================
                if "Halls" in wb.sheetnames:
                    sheet = wb["Halls"]
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        hall_no, building_name, total_seats = row

                        building = Building.objects.get(
                            name=building_name.strip()
                        )

                        ExamHall.objects.update_or_create(
                            hall_no=int(hall_no),
                            defaults={
                                "building": building,
                                "total_seats": int(total_seats)
                            }
                        )

            messages.success(request, "Master data uploaded successfully!")

        except Exception as e:
            messages.error(request, f"Upload failed: {str(e)}")

        return redirect("dashboard")

    return render(request, "allotment/master_upload.html")



@never_cache
@login_required(login_url="login")
def seating_overview(request):

    query = request.GET.get("q")
    hall_filter = request.GET.get("hall")
    session_filter = request.GET.get("session")

    allotments = SeatingAllotment.objects.select_related(
        'student',
        'student__department',
        'exam',
        'exam__subject',
        'hall',
        'hall__building'
    )

    # 🔎 SEARCH FILTER
    if query:
        allotments = allotments.filter(
            Q(student__name__icontains=query) |
            Q(student__reg_no__icontains=query) |
            Q(student__department__name__icontains=query) |
            Q(exam__subject__subject_name__icontains=query)
        )

    # 🏢 HALL FILTER
    if hall_filter and hall_filter != "all":
        allotments = allotments.filter(hall__id=hall_filter)

    # ⏰ SESSION FILTER
    if session_filter and session_filter != "all":
        allotments = allotments.filter(exam__session=session_filter)

    allotments = allotments.order_by(
        'exam__exam_date',
        'exam__session',
        'hall__hall_no',
        'seat_number'
    )

    from .models import ExamHall
    halls = ExamHall.objects.all()

    return render(request, "allotment/seating_overview.html", {
        "allotments": allotments,
        "halls": halls
    })


@never_cache
@login_required(login_url="login") 
def export_master_pdf(request):

    allotments = SeatingAllotment.objects.select_related(
        'student',
        'student__department',
        'exam',
        'exam__subject',
        'hall'
    ).annotate(
        session_order=Case(
            When(exam__session='FN', then=0),
            When(exam__session='AN', then=1),
            output_field=IntegerField()
        )
    ).order_by(
        'exam__exam_date',
        'session_order',
        'hall__hall_no',
        'seat_number',
        'student__department__name'
    )

    html_string = render_to_string(
        'allotment/print.html',
        {
            'allotments': allotments,
            'now': now()
        }
    )

    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Seating_Master_List.pdf"'

    return response

@never_cache
@login_required(login_url="login")
def download_master_template(request):

    wb = Workbook()

    # ---------------------------
    # Departments Sheet
    # ---------------------------
    sheet = wb.active
    sheet.title = "Departments"
    sheet.append(["Department Name"])
    sheet.append(["Computer Science"])
    sheet.append(["Commerce"])

    # ---------------------------
    # Students Sheet
    # ---------------------------
    sheet = wb.create_sheet("Students")
    sheet.append(["Reg No", "Name", "Department Name", "Semester"])
    sheet.append(["23CS001", "John Doe", "Computer Science", 1])

    # ---------------------------
    # Subjects Sheet
    # ---------------------------
    sheet = wb.create_sheet("Subjects")
    sheet.append(["Subject Code", "Subject Name", "Department Name", "Semester"])
    sheet.append(["CS101", "Data Structures", "Computer Science", 1])

    # ---------------------------
    # Exams Sheet
    # ---------------------------
    sheet = wb.create_sheet("Exams")
    sheet.append(["Subject Code", "Exam Date (YYYY-MM-DD)", "Session (FN/AN)"])
    sheet.append(["CS101", "2026-03-15", "FN"])

    # ---------------------------
    # Buildings Sheet
    # ---------------------------
    sheet = wb.create_sheet("Buildings")
    sheet.append(["Building Name"])
    sheet.append(["Main Block"])

    # ---------------------------
    # Halls Sheet
    # ---------------------------
    sheet = wb.create_sheet("Halls")
    sheet.append(["Hall No", "Building Name", "Total Seats"])
    sheet.append([101, "Main Block", 40])

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=Master_Template.xlsx"

    wb.save(response)
    return response